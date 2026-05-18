#!/usr/bin/env python3
"""Diagnóstico del Excel — muestra por qué faltan filas"""
import sys, os

def diagnosticar(ruta):
    try:
        import openpyxl
    except ImportError:
        os.system(f"{sys.executable} -m pip install openpyxl --quiet")
        import openpyxl

    print(f"\n📂 Analizando: {ruta}\n")
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb.active
    print(f"📄 Hoja activa: '{ws.title}'")
    max_row = ws.max_row
    if max_row:
        print(f"📏 Filas reportadas por Excel: {max_row:,}")
    else:
        print(f"📏 Filas reportadas por Excel: (no disponible en este archivo)")
    print(f"📐 Columnas reportadas: {ws.max_column}\n")

    total = 0
    vacias_a = 0
    duplicados = 0
    seriales_vistos = set()
    primera = True
    encabezado_saltado = False
    ejemplos_vacios = []
    ejemplos_dup = []

    for i, fila in enumerate(ws.iter_rows(values_only=True), start=1):
        if primera:
            primera = False
            val = str(fila[0] or '').strip()
            if not val.lstrip('-').replace('.','',1).isdigit():
                encabezado_saltado = True
                print(f"ℹ️  Fila 1 detectada como encabezado: {list(fila[:3])}")
                continue

        col_a = fila[0]

        if col_a is None or str(col_a).strip() == '':
            vacias_a += 1
            if len(ejemplos_vacios) < 3:
                ejemplos_vacios.append(f"  fila {i}: {list(fila[:3])}")
            continue

        serial = str(col_a).strip()

        if serial in seriales_vistos:
            duplicados += 1
            if len(ejemplos_dup) < 3:
                ejemplos_dup.append(f"  fila {i}: serial '{serial}'")
        else:
            seriales_vistos.add(serial)

        total += 1

    wb.close()

    print(f"{'─'*45}")
    print(f"✅ Registros válidos exportados : {total:>8,}")
    print(f"⬜ Filas con columna A vacía    : {vacias_a:>8,}")
    print(f"🔁 Seriales duplicados (omitidos): {duplicados:>8,}")
    salto = 1 if encabezado_saltado else 0
    print(f"{'─'*45}")
    print(f"📊 Total filas procesadas        : {total+vacias_a+duplicados+salto:>8,}")

    if ejemplos_vacios:
        print(f"\nEjemplos de filas vacías en col A:")
        print('\n'.join(ejemplos_vacios))
    if ejemplos_dup:
        print(f"\nEjemplos de seriales duplicados:")
        print('\n'.join(ejemplos_dup))

    print()

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Uso: python diagnostico.py archivo.xlsx")
    else:
        diagnosticar(sys.argv[1])
